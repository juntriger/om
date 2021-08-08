from django.http import HttpResponse
from ..models import Order, Process
import openpyxl
from tqdm import tqdm

def export_order_xlsx(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="orderprocess.xlsx"'

    wb=openpyxl.Workbook()
    ws=wb.create_sheet('process')

    columns = [
        'Order Id',
        'Seq#',
        'PO Number',
        'Customer',
        'Type',
        'Date',
        'RTD',
        'ETD',
        'Brand',
        'Item',
        'Color',
        'Pattern',
        'Specification',
        'Order Q\'ty',
        'Unit',
        'Q\'ty',
        'Stock',
        'State',
        'Update',
        'Date_Plan',
        'P/D Date',
        'P/D Qty',
        'B/F Date',
        'B/F Qty',
        'E/B Date',
        'E/B Qty',
        'P/T Date',
        'P/T Qty',
        'I/S Date',
        'I/S Qty',
        'S/H Date',
        'S/H Qty',
        'Model Name',
        'Sample Step',
        'Material Type',
        'Remark'
    ]

    for col_num in range(len(columns)):
        ws.cell(row=1, column=col_num+1).value = columns[col_num]


    order_data = Order.objects.all().values_list(
        'order_id',
        'seq_num',
        'customer_po',
        'customer',
        'order_type',
        'order_date',
        'rtd',
        'etd',
        'brand',
        'item',
        'color_code',
        'pattern',
        'spec',
        'order_qty',
        'unit',
        'model_name',
        'sample_step',
        'material_type',
        'remark',
        'id'
    )

    prcs_list = Process.objects.all().values_list(
        'order',
        'process_qty',
        'procedure',
        'last_state',
        'last_update',
        'plan_date',
        'pd_date',
        'pd_qty',
        'bf_date',
        'bf_qty',
        'eb_date',
        'eb_qty',
        'pt_date',
        'pt_qty',
        'ins_date',
        'ins_qty',
        'shp_date',
        'shp_qty',
    )

    row_num = 2
    # 전체 오더 리스트에서 오더를 불러온다
    for order in tqdm(order_data):
        # 특정 오더의 해당 프로세스를 불러온다
        order_prcs_all = prcs_list.filter(order=order[19])

        for order_prcs in order_prcs_all:
            pcol_num = 15
            ocol_num = 0
            for data in range(1, len(order_prcs)):
                pcol_num = pcol_num + 1
                ws.cell(row=row_num, column=pcol_num).value = order_prcs[data]

            # 오더 정보를 열 마다 기록한다
            for info in range(len(order) - 5):
                ocol_num = ocol_num + 1
                ws.cell(row=row_num, column=ocol_num).value = order[info]

            for info in range(len(order) - 5, len(order)-1):
                pcol_num = pcol_num + 1
                ws.cell(row=row_num, column=pcol_num).value = order[info]

            row_num = row_num + 1

    wb.save(response)

    return response