from opman.models import Order, Process, Customer
from django.utils import timezone
import openpyxl
from tqdm import trange, tqdm
from django.shortcuts import get_object_or_404
import csv

def divide_colorcode(color):
    color_code_list = color.split(' ')
    count_color_code_unit = len(color_code_list)

    if (count_color_code_unit == 1):
        color_code = color

    elif (count_color_code_unit == 2):
        if (color_code_list[1].isdigit()):
            color_code = color_code_list[0] + color_code_list[1]
        elif (')' in color_code_list[1]):
            color_code = color_code_list[0] + color_code_list[1]
        elif (color_code_list[0] == 'FL' or color_code_list[0] == 'PUMA'):
            color_code = color_code_list[0] + ' ' + color_code_list[1]
        else:
            color_code = color_code_list[0]

    elif (count_color_code_unit == 3):
        if (color_code_list[1].isdigit()):
            color_code = color_code_list[0] + color_code_list[1]
        elif (')' in color_code_list[1]):
            color_code = color_code_list[0] + color_code_list[1]
        elif (color_code_list[2].isdigit()):
            color_code = color_code_list[0] + color_code_list[1] + color_code_list[2]
        elif (color_code_list[0] == 'FL' or color_code_list[0] == 'PUMA'):
            color_code = color_code_list[0] + ' ' + color_code_list[1]
        else:
            color_code = color_code_list[0]

    else:
        if (color_code_list[1].isdigit()):
            color_code = color_code_list[0] + color_code_list[1]
        elif (')' in color_code_list[1]):
            color_code = color_code_list[0] + color_code_list[1]
        elif (color_code_list[2].isdigit()):
            color_code = color_code_list[0] + color_code_list[1] + color_code_list[2]
        elif (color_code_list[0] == 'FL' or color_code_list[0] == 'PUMA'):
            color_code = color_code_list[0] + ' ' + color_code_list[1]
        else:
            color_code = color_code_list[0]

    return color_code

def update_by_pending_list():
    file_name = 'C:/Users/juntr/OneDrive/Python Scripts/om/opman/reports/PENDING ORDER.xlsx'
    wb = openpyxl.load_workbook(file_name)
    # ws = wb.get_active_sheet()
    ws = wb.get_sheet_by_name("Sheet1")

    last_row = ws.max_row

    for i in trange(2, last_row+1):

        # 아이템 이름에 후도 결합
        if (int(str(ws[f'K{i}'].value)[3:4]) > 0):
            item = str(ws[f'H{i}'].value) + ' ' + str(ws[f'K{i}'].value)[0:4]
        else:
            item = str(ws[f'H{i}'].value) + ' ' + str(ws[f'K{i}'].value)[0:3]

        # 컬러 이름에서 코드만 추출
        color = divide_colorcode(str(ws[f'J{i}'].value))

        o = Order(
        order_id = str(ws[f'G{i}'].value)[:10],                                  # Order Id
        seq_num = str(ws[f'G{i}'].value)[11:len(str(ws[f'G{i}'].value))],   # Sequence Number
        customer_po = ws[f'E{i}'].value,                                    # PO Number
        customer = ws[f'D{i}'].value,                                       # Customer
        order_type = ws[f'A{i}'].value,                                     # Order Type
        order_date = ws[f'B{i}'].value,                                     # Order Date
        rtd = ws[f'AD{i}'].value,                                           # RTD
        etd = ws[f'AE{i}'].value,                                           # ETD
        brand = str(ws[f'Y{i}'].value[0:3]).upper(),                             # Brand
        item = item,                                                        # Item
        color_code = color,                                                 # Color Code
        pattern = ws[f'I{i}'].value,                                        # Pattern
        spec = ws[f'K{i}'].value,                                           # Specification
        order_qty = ws[f'L{i}'].value,                                      # Ordery Q'ty
        unit = ws[f'Q{i}'].value,                                           # Unit
        price = ws[f'R{i}'].value,                                          # Price($)
        model_name = ws[f'V{i}'].value,                                     # Model Name
        sample_step = ws[f'U{i}'].value,                                    # Sample Step
        material_type = ws[f'AH{i}'].value,                                 # Material Type
        remark = ws[f'AI{i}'].value,                                        # Remark
        modify_date = timezone.now()                                        # Modified Date
        )
        o.save()

def update_new_order():
    file_name = 'C:/Users/juntr/OneDrive/Python Scripts/om/opman/reports/BSVRECIPT.xlsx'
    wb = openpyxl.load_workbook(file_name)
    # ws = wb.get_active_sheet()
    ws = wb.get_sheet_by_name("total received today")
    last_row = ws.max_row

    for i in trange(2, last_row+1):
        s = None

        order_id = ws[f'B{i}'].value
        seq_num = ws[f'C{i}'].value
        order_qty = ws[f'P{i}'].value

        try:
            s = Order.objects.get(order_id=order_id, seq_num=seq_num)
        except:
            pass

        #
        if s == None and order_qty > 0:
            n = Order()
            n.order_id = ws[f'B{i}'].value
            n.seq_num = ws[f'C{i}'].value
            n.customer_po = ws[f'D{i}'].value
            n.customer = ws[f'E{i}'].value
            n.order_type = ws[f'F{i}'].value
            n.order_date = str(ws[f'G{i}'].value)[0:10]
            n.rtd = ws[f'H{i}'].value
            n.etd = ws[f'I{i}'].value
            n.brand = str(ws[f'J{i}'].value[:3]).upper()

            if (int(str(ws[f'O{i}'].value)[3:4]) > 0):
                n.item = str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:4]
            else:
                n.item = str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:3]

            n.color_code = ws[f'L{i}'].value
            n.pattern = ws[f'M{i}'].value
            n.spec = ws[f'O{i}'].value
            n.order_qty = ws[f'P{i}'].value
            n.unit = ws[f'Q{i}'].value
            n.price = ws[f'R{i}'].value
            n.model_name = ws[f'U{i}'].value
            n.sample_step = ws[f'V{i}'].value
            n.material_type = ws[f'X{i}'].value
            n.remark = ws[f'T{i}'].value

            n.save()

        # 기존 오더 제거
        elif s != None and order_qty < 0:
            s.state = False
            s.save()

        # 수정된 오더 재 입력 시
        elif order_qty > 0:

            modify_log = ""
            modify_log = modify_log + str(f"[{str(timezone.now())[0:10]}]: ")

            if s.customer != ws[f'E{i}'].value:
                modify_log = modify_log + str(f"{s.customer}>>>{ws[f'E{i}'].value},")

            if s.order_type != ws[f'F{i}'].value:
                modify_log = modify_log + str(f"{s.order_type}>>>{ws[f'F{i}'].value},")

            if s.rtd != ws[f'H{i}'].value:
                modify_log = modify_log + str(f"{s.rtd}>>>{str(ws[f'H{i}'].value)[0:10]},")

            if s.etd != ws[f'I{i}'].value:
                modify_log = modify_log + str(f"{s.etd}>>>{str(ws[f'I{i}'].value)[0:10]},")

            if s.color_code != ws[f'L{i}'].value:
                modify_log = modify_log + str(f"{s.color_code}>>>{ws[f'L{i}'].value},")

            if s.pattern != ws[f'M{i}'].value:
                modify_log = modify_log + str(f"{s.pattern}>>>{ws[f'M{i}'].value},")

            if s.spec != ws[f'O{i}'].value:
                modify_log = modify_log + str(f"{s.spec}>>>{ws[f'O{i}'].value},")

            if s.order_qty != ws[f'P{i}'].value:
                modify_log = modify_log + str(f"{s.order_qty}M>>>{ws[f'P{i}'].value}M,")

            if s.remark != ws[f'T{i}'].value:
                modify_log = modify_log + str(f"{s.remark}>>>{ws[f'T{i}'].value},")

            s.customer_po = ws[f'D{i}'].value
            s.customer = ws[f'E{i}'].value
            s.order_type = ws[f'F{i}'].value
            s.rtd = ws[f'H{i}'].value
            s.etd = ws[f'I{i}'].value
            s.brand = ws[f'J{i}'].value

            if (int(str(ws[f'O{i}'].value)[3:4]) > 0):
                if s.item != str(str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:4]):
                    modify_log = modify_log + str(f"{s.item}>>>{str(str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:4])},")
                s.item = str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:4]
            else:
                if s.item != str(str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:3]):
                    modify_log = modify_log + str(f"{s.item}>>>{str(str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:3])},")
                s.item = str(ws[f'K{i}'].value) + ' ' + str(ws[f'O{i}'].value)[0:3]

            s.color_code = ws[f'L{i}'].value
            s.pattern = ws[f'M{i}'].value
            s.spec = ws[f'O{i}'].value
            s.order_qty = ws[f'P{i}'].value
            s.price = ws[f'R{i}'].value
            s.remark = ws[f'T{i}'].value
            s.model_name = ws[f'U{i}'].value
            s.sample_step = ws[f'V{i}'].value
            s.modify_date = timezone.now()
            s.modify_log = modify_log

            s.state = ''
            s.save()

        else:
            pass

def update_customer_list():
    file_name = 'C:/Users/juntr/OneDrive/Python Scripts/om/opman/reports/customer_list.xlsx'
    wb = openpyxl.load_workbook(file_name)
    ws = wb.get_sheet_by_name("customer_list")
    last_row = ws.max_row

    for i in trange(1, last_row+1):
        c = Customer(
            customer=ws[f'A{i}'].value
        )
        c.save()

def find_date_in_cell(cell, str):

    if len(cell[0:cell.upper().find(str)].split('/')[0]) == 1:
        mm = '0' + cell[0:cell.upper().find(str)].split('/')[0]
    else:
        mm = cell[0:cell.upper().find(str)].split('/')[0]

    if len(cell[0:cell.upper().find(str)].split('/')[1]) == 1:
        dd = '0' + cell[0:cell.upper().find(str)].split('/')[1]
    else:
        dd = cell[0:cell.upper().find(str)].split('/')[1]

    date = f"2021-{mm}-{dd}"

    return date


def arrange_process(process_lst):
    """
    A : addition
    C : complete
    S : shortage
    B : B grade
    P : prepare
    W : wait
    E : error
    """
    procedure \
        = plan_state \
        = state_last \
        = date_last \
        = pd_state \
        = bf_state \
        = eb_state \
        = pt_state \
        = ins_state \
        = None

    procedure = str(process_lst[0])
    plan_date = str(process_lst[1])
    pd_qty = process_lst[2]
    pd_date = str(process_lst[3])
    bf_qty = process_lst[4]
    bf_date = str(process_lst[5])
    eb_qty = process_lst[6]
    eb_date = str(process_lst[7])
    pt_qty = process_lst[8]
    pt_date = str(process_lst[9])
    ins_qty = process_lst[10]
    ins_date = str(process_lst[11])
    shp_qty = process_lst[12]
    shp_date = str(process_lst[13])

    if procedure == 'N':
        procedure = 'Production'
    else:
        procedure = 'Stock'

    process = {
        'procedure': procedure,
        'first_update': None,
        'last_update': None,
        'last': {
            'prior': 8,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'plan': {
            'prior': 7,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'pd': {
            'prior': 6,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'bf': {
            'prior': 5,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'eb': {
            'prior': 4,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'pt': {
            'prior': 3,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'ins': {
            'prior': 2,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        },
        'shp': {
            'prior': 1,
            'state': None,
            'date': None,
            'qty': None,
            'annotation': None
        }
    }

    # Plan 열
    if plan_date == "*":
        pass

    elif plan_date.upper().find(",S") > 0:
        process['last']['state'] = "S"
        process['last']['date'] = find_date_in_cell(plan_date, ",S")

    elif plan_date.upper().find(",W") > 0:
        process['last']['state'] = "E"
        process['last']['date'] = find_date_in_cell(plan_date, ",W")
        process['plan']['state'] = "C"
        process['plan']['date'] = find_date_in_cell(plan_date, ",W")

    elif plan_date.upper().find(",C") > 0:
        process['last']['state'] = "S"
        process['last']['date'] = find_date_in_cell(plan_date, ",C")
        process['plan']['state'] = "A"
        process['plan']['date'] = find_date_in_cell(plan_date, ",C")

    elif plan_date.upper().find(",B") > 0:
        process['last']['state'] = "B"
        process['last']['date'] = find_date_in_cell(plan_date, ",B")

    elif plan_date.upper().find(" KH GAP") > 0:
        process['plan']['state'] = "C"
        process['plan']['date'] = find_date_in_cell(plan_date, " KH GAP")

    elif plan_date.upper().find("QC LAB") > 0 :
        process['procedure'] = "LAB"

    elif plan_date.upper().find(" REPRODUCE") > 0:
        process['plan']['state'] = "C"
        process['plan']['date'] = find_date_in_cell(plan_date, " REPRODUCE")

    elif plan_date.upper().find(" RECOATING") > 0:
        process['plan']['state'] = "C"
        process['plan']['date'] = find_date_in_cell(plan_date, " RECOATING")

    elif plan_date.upper().find(" HOLD") > 0:
        process['plan']['state'] = "H"
        process['plan']['date'] = find_date_in_cell(plan_date, " HOLD")
    else:
        pass

    # Production Qty 열
    if type(pd_qty) == int:
        process['pd']['state'] = "C"
        process['pd']['date'] = pd_date[:10]
        process['pd']['qty'] = pd_qty

    elif pd_qty == "*" or pd_qty == None:
        pass

    elif pd_qty.upper().find(",B") > 0:
        process['last']['state'] = "B"
        process['last']['date'] = find_date_in_cell(pd_qty, ",B")

    elif pd_qty.upper().find(",S") > 0:
        process['last']['state'] = "S"
        process['last']['date'] = find_date_in_cell(pd_qty, ",S")

    elif pd_qty.upper().find(",C") > 0:
        process['last']['state'] = "S"
        process['last']['date'] = find_date_in_cell(pd_qty, ",C")
        process['pd']['state'] = "A"
        process['pd']['date'] = find_date_in_cell(pd_qty, ",C")
    else:
        pass

    # 버핑 열
    if type(bf_qty) == int:
        process['bf']['state'] = "C"
        process['bf']['date'] = bf_date[:10]
        process['bf']['qty'] = bf_qty
    else:
        pass

    # 앰보 열
    if type(eb_qty) == int:
        process['eb']['state'] = "C"
        process['eb']['date'] = eb_date[:10]
        process['eb']['qty'] = eb_qty
    else:
        pass

    # 컬러 처리열
    if type(pt_qty) == int:
        process['pt']['state'] = "C"
        process['pt']['date'] = pt_date[:10]
        process['pt']['qty'] = pt_qty
    else:
        pass

    # 검사 열
    if type(ins_qty) == int:
        process['ins']['state'] = "C"
        process['ins']['date'] = ins_date[:10]
        process['ins']['qty'] = ins_qty
    elif ins_qty == "C/S":
        process['ins']['state'] = "P"
    elif ins_qty == "ST":
        process['ins']['state'] = "P"
    elif ins_qty == "RE-TEST":
        process['ins']['state'] = "W"
    elif ins_qty == "W/bonding":
        process['ins']['state'] = "W"
    else:
        pass

    if type(shp_qty) == int:
        process['shp']['state'] = "Shipped"
        process['shp']['date'] = shp_date[:10]
        process['shp']['qty'] = shp_qty

    if (
            process['last']['date'] or
            process['plan']['date'] or
            process['pd']['date'] or
            process['bf']['date'] or
            process['eb']['date'] or
            process['pt']['date'] or
            process['ins']['date'] or
            process['shp']['date']
    ):
        process['first_update'] = min(list(filter(None,[process['last']['date'],
                                      process['plan']['date'],
                                      process['pd']['date'],
                                      process['bf']['date'],
                                      process['eb']['date'],
                                      process['pt']['date'],
                                      process['ins']['date'],
                                      process['shp']['date']
                                      ])))

        process['last_update'] = max(list(filter(None,[process['last']['date'],
                                      process['plan']['date'],
                                      process['pd']['date'],
                                      process['bf']['date'],
                                      process['eb']['date'],
                                      process['pt']['date'],
                                      process['ins']['date'],
                                      process['shp']['date']
                                      ])))

    return process

def opfile_to_dict(file_name):

    """
    오더 프로세스 파일의 프로세스 부분을
    딕셔너리로 변환하는 함수
    """

    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb.get_sheet_by_name("Process Order")
    last_row = ws.max_row

    opsdict = {}
    for row in trange(2, last_row + 1):
        process_lst = [
            ws[f'S{row}'].value,
            ws[f'T{row}'].value,
            ws[f'U{row}'].value,
            ws[f'V{row}'].value,
            ws[f'W{row}'].value,
            ws[f'X{row}'].value,
            ws[f'Y{row}'].value,
            ws[f'Z{row}'].value,
            ws[f'AA{row}'].value,
            ws[f'AB{row}'].value,
            ws[f'AC{row}'].value,
            ws[f'AD{row}'].value,
            ws[f'AE{row}'].value,
            ws[f'AF{row}'].value,
        ]
        order_key = ws[f'A{row}'].value + '-' + str(ws[f'B{row}'].value)
        pdict = {**{'process_qty': ws[f'O{row}'].value}, **arrange_process(process_lst)}

        if order_key in opsdict:
            process_num = len(opsdict[order_key])
            opsdict[order_key] = {**opsdict[order_key], **{process_num: pdict}}
            opsdict.update(opsdict)
        else:
            process_num = 0
            opdict = {order_key: {process_num: pdict}}
            opsdict = {**opsdict, **opdict}

        #with open('opsdict.csv', 'w') as f:
        #    w = csv.writer(f)
        #    w.writerow(opsdict.keys())
        #    w.writerow(opsdict.values())

    return opsdict

def compare_process_dict(dic_a, dic_b):

    """
    두 딕셔너리 비교 함수
    dic_a : DB에 등록된 프로세스 목록
    dic_b : 파일에서 갱신한 프로세스 목록
    """
    prsc = {}
    # 두 딕셔너리 비교
    prcs_qty_a = []
    for prcs_a in dic_a:
        prcs_qty_a = prcs_qty_a.append(dic_a[prcs_a]['process_qty'])

    for prcs_a in dic_a:
        for prcs_b in dic_b:

            # 진행 수량이 동일한 프로세스라면
            if dic_a[prcs_a]['process_qty'] == dic_b[prcs_b]['process_qty']:

                # 최초 업데이트 일자가 같다면
                if dic_a[prcs_a]['first_update'] == dic_b[prcs_b]['first_update']:
                    prsc = {**prsc, **dic_b[prcs_b]}

                # 최초 업데이트 일자가 다른 수량만 같은 프로세스라면
                else:

                    # 기존 프로세스가 최초 업데이트 일자가 없었던 것이라면
                    if dic_a[prcs_a]['first_update'] == None:
                        prsc = {**prsc, **dic_b[prcs_b]}

                    # 수량만 같은 추가 프로세스
                    else:
                        prsc = {**prsc, **dic_a[prcs_a], **dic_b[prcs_b]}

            # 진행 수량이 다른, 추가 진행 프로세스라면
            else:
                if dic_b[prcs_b]['process_qty'] in prcs_qty_a:
                    pass
                else:
                    prsc = {**prsc, **dic_a[prcs_a], **dic_b[prcs_b]}
    
    return prsc

def qset_to_dict(qset):
    dict = {}
    i = 0
    for q in qset:
        process = {
            i:{
                'procedure': q.procedure,
                'process_qty': q.process_qty,
                'first_update': q.first_update,
                'last_update': q.last_update,
                'last': {
                    'prior': 8,
                    'state': q.state_last,
                    'date': q.date_last,
                    'qty': None,
                    'annotation': None
                },
                'plan': {
                    'prior': 7,
                    'state': q.plan_state,
                    'date': q.plan_date,
                    'qty': q.plan_qty,
                    'annotation': None
                },
                'pd': {
                    'prior': 6,
                    'state': q.pd_state,
                    'date': q.pd_date,
                    'qty': q.pd_qty,
                    'annotation': None
                },
                'bf': {
                    'prior': 5,
                    'state': q.bf_state,
                    'date': q.bf_date,
                    'qty': q.bf_qty,
                    'annotation': None
                },
                'eb': {
                    'prior': 4,
                    'state': q.eb_state,
                    'date': q.eb_date,
                    'qty': q.eb_qty,
                    'annotation': None
                },
                'pt': {
                    'prior': 3,
                    'state': q.pt_state,
                    'date': q.pt_date,
                    'qty': q.pt_qty,
                    'annotation': None
                },
                'ins': {
                    'prior': 2,
                    'state': q.ins_state,
                    'date': q.ins_date,
                    'qty': q.ins_qty,
                    'annotation': None
                },
                'shp': {
                    'prior': 1,
                    'state': None,
                    'date': q.shp_date,
                    'qty': q.shp_qty,
                    'annotation': None
                }
            }
        }
        dict = {**dict, **process}
        i = i + 1

    return dict

def update_by_orderprocess():

    file_name = 'C:/Users/juntr/OneDrive/Python Scripts/om/opman/reports/orderprocess.xlsx'

    # 오더 프로세스 딕셔너리화
    opsdict = opfile_to_dict(file_name)

    order_list = Order.objects.filter(state=None)

    # Order 모델의 order 가 주체가 되어 opsdict 검색
    for order in tqdm(order_list):
        try:
            order_key = order.order_id + '-' + str(order.seq_num)
            dict_b = opsdict[order_key]

            if Process.objects.filter(order=order.id):    # opsdict에 오더 키값이 있다면
                dict_a = qset_to_dict(Process.objects.filter(order=order.id))
                prsc_dict = compare_process_dict(dict_a, dict_b)

            else:
                prsc_dict = dict_b

            # 딕셔너리 포함된 모든 프로세스 업데이트
            for prsc in prsc_dict:
                p = Process(
                    order=order,
                    process_qty=prsc_dict[prsc]['process_qty'],
                    procedure=prsc_dict[prsc]['procedure'],  # Production Procedure
                    first_update=prsc_dict[prsc]['first_update'],
                    last_state=None,  # Status
                    last_update=prsc_dict[prsc]['last_update'],  # Update Date
                    plan_state=prsc_dict[prsc]['plan']['state'],
                    plan_date=prsc_dict[prsc]['plan']['date'],  # Plan Date
                    plan_qty = None,  # Plan Q'ty
                    state_last=prsc_dict[prsc]['last']['state'],
                    date_last=prsc_dict[prsc]['last']['date'],
                    pd_state=prsc_dict[prsc]['pd']['state'],
                    pd_date=prsc_dict[prsc]['pd']['date'],  # Production Date
                    pd_qty=prsc_dict[prsc]['pd']['qty'],  # Production Q'ty
                    bf_state=prsc_dict[prsc]['bf']['state'],
                    bf_date=prsc_dict[prsc]['bf']['date'],  # Buffing Date
                    bf_qty=prsc_dict[prsc]['bf']['qty'],  # Buffing Q'ty
                    eb_state=prsc_dict[prsc]['eb']['state'],
                    eb_date=prsc_dict[prsc]['eb']['date'],  # Embo Date
                    eb_qty=prsc_dict[prsc]['eb']['qty'],  # Embo Q'ty
                    pt_state=prsc_dict[prsc]['pt']['state'],
                    pt_date=prsc_dict[prsc]['pt']['date'],  # Color Print Date
                    pt_qty=prsc_dict[prsc]['pt']['qty'],  # Color Print Q'ty
                    ins_state=prsc_dict[prsc]['ins']['state'],
                    ins_date=prsc_dict[prsc]['ins']['date'],  # Inspection Date
                    ins_qty=prsc_dict[prsc]['ins']['qty'],  # Inspection Q'ty
                    shp_date=prsc_dict[prsc]['shp']['date'],  # Shipment Date
                    shp_qty=prsc_dict[prsc]['shp']['qty']
                )
                p.save()
        except:
            pass


    """
    # row에 해당되는 오더가 이미 있다면
    if order.id:

        # 기존과 동일한 지 확인
        process_old = Process.objects.get(pk=order.id)
        if (
                (process_old.process_qty == ws[f'O{row}'].value) and  # 진행 수량 동일
                (process_old.state == state) and                    # 진행 상태 동일
                (process_old.update == update)                      # 업데이트 날짜 동일
        ):
            pass
        # 동일하지 않다면
        else:
            # 동일 프로세스 업데이트인지
            # 추가 프로세스 진행 정보인지
            pass

    # 해당되는 오더가 없다면
    else:
        
    """
    return